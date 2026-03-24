import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws1 = wb.active
ws1.title = "基础数据"
ws2 = wb.create_sheet("月度推算表")
ws3 = wb.create_sheet("备料投产节点")

# ==================== 样式定义 ====================
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=12)
warn_fill = PatternFill("solid", fgColor="FFFF99")   # 黄色：超产能月份
cap_fill  = PatternFill("solid", fgColor="FFD9B3")   # 橙色：满产月份
head_fill = PatternFill("solid", fgColor="BDD7EE")   # 蓝色：表头
s1_fill   = PatternFill("solid", fgColor="E2EFDA")   # 绿色：Sheet1表头
center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
left      = Alignment(horizontal="left",  vertical="center", wrap_text=True)

thin = Side(style="thin")
def all_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, fill=head_fill):
    cell.font = header_font
    cell.fill = fill
    cell.alignment = center
    cell.border = all_border()

def style_data(cell, fill=None):
    cell.alignment = center
    cell.border = all_border()
    if fill:
        cell.fill = fill

# ==================== SHEET 1: 基础数据 ====================
ws1.merge_cells("A1:D1")
ws1["A1"] = "产销推算表 — 基础数据"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A1"].alignment = center
ws1["A1"].fill = PatternFill("solid", fgColor="4472C4")
ws1["A1"].font = Font(bold=True, size=14, color="FFFFFF")

headers1 = ["参数", "黑色", "银色", "说明"]
for c, h in enumerate(headers1, 1):
    cell = ws1.cell(row=2, column=c, value=h)
    style_header(cell, s1_fill)

s1_rows = [
    ("初始库存",         989,  1004, "3月初期初库存"),
    ("在产入库（3月）",  1400,  600, "2月已投产，3月到库（含在该行）"),
    ("月产能上限（黑+银）", 3500, "/", "每月黑色+银色总投产 ≤ 3500"),
]
for r, (a, b, c, d) in enumerate(s1_rows, 3):
    for col, val in enumerate([a, b, c, d], 1):
        cell = ws1.cell(row=r, column=col, value=val)
        style_data(cell)

ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 10
ws1.column_dimensions["C"].width = 10
ws1.column_dimensions["D"].width = 36

# ==================== SHEET 2: 月度推算表 ====================
# 列: A月份 B期初黑 C期初银 D在产入库黑 E在产入库银 F本月投产黑 G本月投产银 H总投产 I需求黑 J需求银 K期末黑 L期末银 M备料月份 N投产月份

ws2.merge_cells("A1:N1")
ws2["A1"] = "月度产销推算表（期末库存 = 期初库存 + 在产入库 + 本月投产 − 本月需求）"
ws2["A1"].font = Font(bold=True, size=12, color="FFFFFF")
ws2["A1"].fill = PatternFill("solid", fgColor="4472C4")
ws2["A1"].alignment = center

headers2 = [
    "月份",
    "期初库存\n黑色", "期初库存\n银色",
    "在产入库\n黑色", "在产入库\n银色",
    "本月投产\n黑色", "本月投产\n银色",
    "总投产",
    "需求\n黑色", "需求\n银色",
    "期末库存\n黑色", "期末库存\n银色",
    "备料月份", "投产月份"
]
for c, h in enumerate(headers2, 1):
    cell = ws2.cell(row=2, column=c, value=h)
    style_header(cell)

ws2.row_dimensions[2].height = 36

# 数据: (月份, 在产入库黑, 在产入库银, 本月投产黑, 本月投产银, 需求黑, 需求银, 备料月, 投产月)
months = [
    ("3月",  1400, 600,  0,    0,    0,    0,    "1月",  "2月"),
    ("4月",  0,    0,    0,    0,    140,  60,   "2月",  "3月"),
    ("5月",  0,    0,    0,    0,    182,  78,   "3月",  "4月"),
    ("6月",  0,    0,    0,    0,    676,  310,  "4月",  "5月"),
    ("7月",  0,    0,    93,   40,   1113, 477,  "5月",  "6月"),
    ("8月",  0,    0,    2842, 658,  1638, 702,  "6月",  "7月"),
    ("9月",  0,    0,    2450, 1050, 1918, 822,  "7月",  "8月"),
    ("10月", 0,    0,    2450, 1050, 2996, 1284, "8月",  "9月"),
    ("11月", 0,    0,    2450, 1050, 3829, 1641, "9月",  "10月"),
    ("12月", 0,    0,    2450, 1050, 2632, 1128, "10月", "11月"),
]

# 超产能月份（10-12月）背景色
over_cap_months = {"10月", "11月", "12月"}
full_cap_months = {"8月", "9月"}

for i, (month, zb, zs, pb, ps, db, ds, blyf, tpyf) in enumerate(months):
    row = i + 3  # 数据从第3行开始

    is_over  = month in over_cap_months
    is_full  = month in full_cap_months
    row_fill = warn_fill if is_over else (cap_fill if is_full else None)

    # A: 月份
    c = ws2.cell(row=row, column=1, value=month)
    style_data(c, row_fill)

    # B: 期初黑
    if row == 3:
        val = "='基础数据'!B3"
    else:
        val = f"=K{row-1}"
    c = ws2.cell(row=row, column=2, value=val)
    style_data(c, row_fill)

    # C: 期初银
    if row == 3:
        val = "='基础数据'!C3"
    else:
        val = f"=L{row-1}"
    c = ws2.cell(row=row, column=3, value=val)
    style_data(c, row_fill)

    # D: 在产入库黑
    if row == 3:
        val = "='基础数据'!B4"
    else:
        val = 0
    c = ws2.cell(row=row, column=4, value=val)
    style_data(c, row_fill)

    # E: 在产入库银
    if row == 3:
        val = "='基础数据'!C4"
    else:
        val = 0
    c = ws2.cell(row=row, column=5, value=val)
    style_data(c, row_fill)

    # F: 本月投产黑
    c = ws2.cell(row=row, column=6, value=pb)
    style_data(c, row_fill)

    # G: 本月投产银
    c = ws2.cell(row=row, column=7, value=ps)
    style_data(c, row_fill)

    # H: 总投产 =F+G
    c = ws2.cell(row=row, column=8, value=f"=F{row}+G{row}")
    style_data(c, row_fill)

    # I: 需求黑
    c = ws2.cell(row=row, column=9, value=db)
    style_data(c, row_fill)

    # J: 需求银
    c = ws2.cell(row=row, column=10, value=ds)
    style_data(c, row_fill)

    # K: 期末黑 =B+D+F-I
    c = ws2.cell(row=row, column=11, value=f"=B{row}+D{row}+F{row}-I{row}")
    style_data(c, row_fill)

    # L: 期末银 =C+E+G-J
    c = ws2.cell(row=row, column=12, value=f"=C{row}+E{row}+G{row}-J{row}")
    style_data(c, row_fill)

    # M: 备料月份
    c = ws2.cell(row=row, column=13, value=blyf)
    style_data(c, row_fill)

    # N: 投产月份
    c = ws2.cell(row=row, column=14, value=tpyf)
    style_data(c, row_fill)

# 合计行
sum_row = len(months) + 3
ws2.cell(row=sum_row, column=1, value="合计").font = header_font
ws2.cell(row=sum_row, column=1).border = all_border()
ws2.cell(row=sum_row, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
ws2.cell(row=sum_row, column=1).alignment = center

for col in [6, 7, 8, 9, 10]:  # 投产和需求列汇总
    cell = ws2.cell(row=sum_row, column=col,
                    value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{sum_row-1})")
    cell.font = header_font
    cell.border = all_border()
    cell.fill = PatternFill("solid", fgColor="D9D9D9")
    cell.alignment = center

# 列宽
col_widths = [8, 12, 12, 12, 12, 12, 12, 10, 10, 10, 12, 12, 10, 10]
for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# 备注说明
note_row = sum_row + 2
ws2.merge_cells(f"A{note_row}:N{note_row}")
note = ws2.cell(row=note_row, column=1,
    value="【颜色说明】橙色=满产月（黑+银=3500）；黄色=超产能月（需求>3500，不足部分由期初库存补足）；本月投产=上月投入生产本月到库，备料月=投产月提前1个月")
note.font = Font(italic=True, size=10, color="444444")
note.alignment = left

# ==================== SHEET 3: 备料投产节点 ====================
ws3.merge_cells("A1:G1")
ws3["A1"] = "备料 & 投产时间节点汇总表"
ws3["A1"].font = Font(bold=True, size=12, color="FFFFFF")
ws3["A1"].fill = PatternFill("solid", fgColor="4472C4")
ws3["A1"].alignment = center

headers3 = ["目标月份", "备料月份", "投产月份", "投产总量", "黑色投产", "银色投产", "备注"]
for c, h in enumerate(headers3, 1):
    cell = ws3.cell(row=2, column=c, value=h)
    style_header(cell)

s3_data = [
    ("3月",  "1月",  "2月",  2000, 1400, 600,
     "在产入库（已在产2月，3月到库）"),
    ("4月",  "2月",  "3月",  0,    0,    0,
     "期初库存充足（黑2389/银1604），无需安排新投产"),
    ("5月",  "3月",  "4月",  0,    0,    0,
     "期初库存充足，无需安排新投产"),
    ("6月",  "4月",  "5月",  0,    0,    0,
     "期初库存充足，无需安排新投产"),
    ("7月",  "5月",  "6月",  133,  93,   40,
     "6月投产133件（黑93/银40），补充后续库存"),
    ("8月",  "6月",  "7月",  3500, 2842, 658,
     "7月满产3500件（黑2842/银658）"),
    ("9月",  "7月",  "8月",  3500, 2450, 1050,
     "8月满产3500件（黑2450/银1050）"),
    ("10月", "8月",  "9月",  3500, 2450, 1050,
     "9月满产3500件；需求4280超产能780件，由期初库存补足（黑546+银234）★"),
    ("11月", "9月",  "10月", 3500, 2450, 1050,
     "10月满产3500件；需求5470超产能1970件，由期初库存补足（黑1379+银591）★"),
    ("12月", "10月", "11月", 3500, 2450, 1050,
     "11月满产3500件；需求3760超产能260件，由期初库存补足（黑182+银78）★"),
]

for r, row_data in enumerate(s3_data, 3):
    month_label = row_data[0]
    is_over = month_label in over_cap_months
    is_full = month_label in full_cap_months
    row_fill = warn_fill if is_over else (cap_fill if is_full else None)

    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        style_data(cell, row_fill)
        if c == 7:
            cell.alignment = left  # 备注左对齐

s3_col_widths = [10, 10, 10, 10, 10, 10, 60]
for i, w in enumerate(s3_col_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

note3_row = len(s3_data) + 4
ws3.merge_cells(f"A{note3_row}:G{note3_row}")
note3 = ws3.cell(row=note3_row, column=1,
    value="★ 超产能月份：投产月按满产3500件安排，缺口由此前滚动累计的库存盈余补足。库存盈余来源于前期需求低于产能时预先建立的安全库存。")
note3.font = Font(italic=True, size=10, color="C00000")
note3.alignment = left

# ==================== 保存 ====================
output_path = "/Users/jackson/.openclaw/workspace/产销推算表.xlsx"
wb.save(output_path)
print("Done:", output_path)
